import os.path
from datetime import *
from tkinter import *
from tkinter import filedialog, messagebox
from ttkbootstrap.constants import *
import ttkbootstrap as tb
import datetime
import openpyxl

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError


SCOPES = ["https://www.googleapis.com/auth/calendar"]
sharedCalender = '[HIDDEN]'

date = []
shifts = []
endList = []
startList = []
dayOff = []
userCell = 'None'


class Shift:
    def __init__(self, hour, minute):
        self.hour = hour
        self.minute = minute

    def timecode(self):
        timeC = str(datetime.time(self.hour, self.minute)) + '-04:00'
        return timeC


def dateCapture():
    filepath = fileVar.get()
    WB = openpyxl.load_workbook(filepath)
    WS = WB['IOA']
    startDate = str(WS['C2'].value).split(' ')
    startDate.pop(-1)
    dateValue = startDate[-1].split('-')
    dateValue = datetime.date(int(dateValue[0]), int(dateValue[1]), int(dateValue[2]))
    startDate = dateValue

    counter = -1

    while counter < 7:
        counter += 1
        nextDay = str(startDate + timedelta(days=counter))
        date.append(nextDay)


def findUser():

    filepath = fileVar.get()
    WB = openpyxl.load_workbook(filepath)
    WS = WB['IOA']

    global userCell

    userinput = str(userVar.get()).casefold().capitalize()

    if userinput != '':
        for row in WS.rows:
            for cell in row:
                try:
                    if userinput in cell.value:
                        column = str(WS.cell(row=cell.row, column=2).column_letter)
                        row = str(WS.cell(row=cell.row, column=2).row)
                        userCell = column + row
                        return userCell
                except (AttributeError, TypeError):
                    continue


def shiftCapture():
    filepath = fileVar.get()
    WB = openpyxl.load_workbook(filepath)
    WS = WB['IOA']
    user = userCell
    userColumn = []
    columns = ['C', 'D', 'E', 'F', 'G', 'H', 'I']
    row = ''

    counter = -1

    for i in user:
        if i.isalpha():
            userColumn.append(i)
    if userColumn[-1] in userColumn:
        row = user.replace(str(userColumn[-1]), '')

    while counter < 6:
        counter += 1
        cell = str(columns[(0+counter)]) + row
        shifts.append(str(WS[cell].value))


def shiftConversion():

    dateCapture()
    shiftCapture()

    startHour = 0
    startMinute = 0
    endHour = 0
    endMinute = 0

    counter = -1

    while counter < 6:
        counter += 1
        userSchedule = str(shifts[0+counter]).upper()
        meridian = []

        if userSchedule[0].isdigit() is False:
            dayOff.append(userSchedule)

        if userSchedule not in dayOff:
            for i in userSchedule:
                if i.isalpha():
                    if i == 'A' or i == 'P':
                        meridian.append(i)
                        if 'P' in meridian:
                            updatedStr = userSchedule.replace('P', '')
                            userSchedule = updatedStr

                        if 'A' in meridian:
                            updatedStr = userSchedule.replace('A', '')
                            userSchedule = updatedStr

        times = userSchedule.split('-')
        start = str(times[0])
        end = str(times[-1])

        if start not in dayOff:
            if ':' in start:
                startSeparate = start.split(':')
                startHour = int(startSeparate[0])
                startMinute = int(startSeparate[-1])
            else:
                startHour = int(start)
                startMinute = 0

            if ':' in end:
                endSeparate = end.split(':')
                endHour = int(endSeparate[0])
                endMinute = int(endSeparate[-1])
            else:
                endHour = int(end)
                endMinute = 0

        if start not in dayOff:
            if startHour > endHour and meridian[-1] == 'P':
                endHour += 12
            elif endHour > startHour and meridian[-1] == 'P':
                endHour += 12
                startHour += 12
                if endHour == 24:
                    endHour -= 24
            elif startHour > endHour and meridian[-1] == 'A':
                startHour += 12

        beginningOfShift = Shift(startHour, startMinute)
        endOfShift = Shift(endHour, endMinute)

        startList.append(beginningOfShift.timecode())
        endList.append(endOfShift.timecode())

        if start in dayOff:
            date.pop(len(startList)-1)
            startList.pop(len(startList)-1)
            endList.pop(len(startList)-1)
            
# This function uses a while loop to convert the lists into Calendar events.

def Upload():
    counter = -1
    stopper = len(endList) - 1
    creds = None
    if os.path.exists("token.json"):
        creds = Credentials.from_authorized_user_file("token.json", SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                "credentials.json", SCOPES
            )
            creds = flow.run_local_server(port=0)
        with open("token.json", "w") as token:
            token.write(creds.to_json())

    while counter < stopper:
        counter += 1
        try:
            service = build("calendar", "v3", credentials=creds)
            test = str(endList[0 + counter])
            if int(test[0]) == 0:
                event = {
                    'summary': "Klonegun's Work Schedule",
                    'start': {'dateTime': str(date[0+counter]) + 'T' + str(startList[0+counter])},
                    'end': {'dateTime': str(date[0+(counter+1)]) + 'T' + str(endList[0+counter])},
                    }

                event = service.events().insert(calendarId=sharedCalender, body=event).execute()
            else:
                event = {
                    'summary': "Klonegun's Work Schedule",
                    'start': {'dateTime': str(date[0 + counter]) + 'T' + str(startList[0 + counter])},
                    'end': {'dateTime': str(date[0 + counter]) + 'T' + str(endList[0 + counter])},
                }

                event = service.events().insert(calendarId=sharedCalender, body=event).execute()

        except HttpError as error:
            print(f"An error occurred {error}")

# GUI Functions and Layout

def submit():
    findUser()
    if userCell == 'None':
        messagebox.showerror(title='ERROR', message='Invalid Entry, please try again.')
        textBox.delete('0', 'end')
    else:
        lastname_Label = tb.Label(root, text='Last name entered: ' + userVar.get().casefold().capitalize(),
                                  font=('Arial', 10))
        lastname_Label.pack(pady=(0, 25))
        run_Button.pack(pady=(0, 25))
        textBox.destroy()
        submit_Button.destroy()
        entry_label.destroy()


def run():
    shiftConversion()
    Upload()
    messagebox.showinfo(title='Success!', message='Your schedule has been successfully uploaded to Google Calendar!')
    quitApp()


def uploadExcel():
    uploadedFile = filedialog.askopenfilename(title="Select a file", filetypes=[("Text files", "*.xlsx"),
                                                                                ("All files", "*.*")])

    if uploadedFile == '':
        messagebox.showerror(title='ERROR', message='Please select an Excel file!')
    else:
        filename_Label = tb.Label(root, text='File Uploaded: ' + uploadedFile, font=('Arial', 10))

        instruct_Label.pack_forget()
        upload_Button.pack_forget()

        filename_Label.pack(side=TOP, pady=25)
        entry_label.pack(pady=(25, 0))
        textBox.pack(pady=(15, 5))
        submit_Button.pack(pady=5)
        fileVar.set(uploadedFile)


def quitApp():
    root.quit()
    root.mainloop()


root = tb.Window(themename='superhero')
root.title('Excel to Google Calendar App')
root.geometry('700x350')

userVar = StringVar()
fileVar = StringVar()

# Label
welcome_Label = tb.Label(root, text='Welcome to the Excel to Google Calendar Converter!', bootstyle='default',
                         font=('Arial', 18, 'bold'))
welcome_Label.pack(pady=25)

entry_label = tb.Label(root, text='Enter your last name below:', font='Arial')

instruct_Label = tb.Label(root, text='Upload your schedule file and follow the prompts.', font='Arial')
instruct_Label.pack(pady=15)

textBox = tb.Entry(root, textvariable=userVar)

# Buttons
upload_Button = tb.Button(root, text='Upload Excel File', command=uploadExcel, bootstyle='info')
upload_Button.pack(pady=35)

submit_Button = tb.Button(root, text='Submit', command=submit, bootstyle='info')

run_Button = tb.Button(root, text='Update Calendar', command=run, bootstyle='info')

quit_Button = tb.Button(root, text='Quit', command=quitApp, bootstyle='danger')
quit_Button.pack(side=BOTTOM, pady=(0, 25))

root.mainloop()
